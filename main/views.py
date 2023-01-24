import datetime
import json

import pandas as pd
import requests
from django.shortcuts import render
from openpyxl import load_workbook

from .models import GeoData, DemandData, SkillsData


def index(request):
    return render(request, 'main.html', {'title': 'aboba'})


def demand(request):
    return render(request, 'demand.html', {'data': DemandData.objects.first()})


def geo(request):
    temp = load_workbook(GeoData.objects.first().docs)
    return render(request, 'geo.html', {'workbook': temp})


def skills(request):
    temp = load_workbook(SkillsData.objects.first().docs)
    return render(request, 'skills.html', {'workbook': temp})


def recent(request):
    recTemp = getHHRUData().values.tolist()[::-1]
    return render(request, 'recent.html', {'recentVacDF': recTemp})


def getHHRUData():
    def get_page(i):
        parameters = {
            "specialization": 1,
            "found": 1,
            "per_page": 100,
            "page": i,
            "order_by": "publication_time",
            "date_from": "2022-12-30T00:00:00+0300",
            "date_to": "2022-12-30T23:59:00+0300"
        }

        try:
            request = requests.get('https://api.hh.ru/vacancies', parameters)
            info = request.content.decode()
            request.close()
        except:
            print("¯\\_(ツ)_/¯")
            return get_page(i)

        print("Успешный запрос")
        return info

    def set_vacancies():
        df = pd.DataFrame(
            columns=['name', 'description', 'key_skills', 'employer', 'salary_mid', 'area_name', 'published_at', 'link', 'employerURL'])
        for i in range(100):
            result = []
            js_obj = json.loads(get_page(i))
            result.append(js_obj)

            if (js_obj['pages'] - i) <= 1:
                break

            for vac in result:
                for r in vac['items']:
                    if any(x.lower() in ['с#', 'c sharp', 'шарп', 'c#'] for x in r['name'].split(' ')):
                        metaSkills, metaDesc = getMeta(r['id'])
                        df.loc[len(df)] = [r['name'], metaDesc,
                                           metaSkills,
                                           r['employer']['name'],
                                           salSet(r),
                                           r['area']['name'],
                                           datetime.datetime.strptime(r['published_at'], "%Y-%m-%dT%H:%M:%S%z"),
                                           'https://hh.ru/vacancy/' + str(r['id']),
                                           r['employer']['alternate_url']
                                           ]
                    if df[df.columns[0]].count() >= 10:
                        return df
        return df

    def getMeta(vacID):
        request = requests.get('https://api.hh.ru/vacancies/' + vacID)
        info = request.content.decode()
        request.close()
        js_obj = json.loads(info)

        metaTempSkills = js_obj['key_skills']
        metaDesc = js_obj['description']
        metaSkills = ''
        for k in metaTempSkills:
            metaSkills += k['name'] + ', '
        metaSkills = metaSkills[:-2]

        return metaSkills, metaDesc

    def salSet(r):
        if r['salary'] is None:
            return 'Не указано'
        elif (r['salary']['from'] is not None) and (r['salary']['to'] is not None):
            return str((r['salary']['from'] + r['salary']['to']) / 2) + ' ' + str(r['salary']['currency'])
        elif (r['salary']['from'] is not None) and (r['salary']['to'] is None):
            return str(r['salary']['from']) + ' ' + str(r['salary']['currency'])
        else:
            return str(r['salary']['to']) + ' ' + str(r['salary']['currency'])

    return set_vacancies()
