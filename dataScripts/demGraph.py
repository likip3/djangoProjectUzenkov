import csv
import math
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

currencyTransfer = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class Vacancy:
    def __init__(self, vacancies):
        self.name = vacancies[name_list[0]]
        self.publicationYear = int(vacancies[name_list[5]][:4])
        self.averageSalary = math.floor((float(vacancies[name_list[1]]) + float(vacancies[name_list[2]])) / 2) * \
                             currencyTransfer[vacancies[name_list[3]]]
        self.areaName = vacancies[name_list[4]]
        self.publicationYear = int(vacancies[name_list[5]][:4])


class DataSet:
    def __init__(self, filename, vacancyName):
        self.filename = filename
        self.vacancyName = vacancyName

    def csv_reader_raw(self):
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            headerLength = len(header)
            for row in reader:
                if '' not in row and len(row) == headerLength:
                    yield dict(zip(header, row))

    @staticmethod
    def increment(dictIN, k, c):
        if k in dictIN:
            dictIN[k] += c
        else:
            dictIN[k] = c

    @staticmethod
    def mean(dictAV):
        newDict = {}
        for k, v in dictAV.items():
            newDict[k] = int(sum(v) / len(v))
        return newDict

    def getDynamics(self):
        salary = {}
        salaryByName = {}
        city = {}
        count = 0

        for vacancyDictionary in self.csv_reader_raw():
            vacancy = Vacancy(vacancyDictionary)
            self.increment(salary, vacancy.publicationYear, [vacancy.averageSalary])
            if any(x.lower() in ['с#', 'c sharp', 'шарп', 'c#'] for x in vacancy.name.split(' ')):
                self.increment(salaryByName, vacancy.publicationYear, [vacancy.averageSalary])
            self.increment(city, vacancy.areaName, [vacancy.averageSalary])
            count += 1

        numberByName = dict([(k, len(v)) for k, v in salaryByName.items()])
        vacancyNum = dict([(k, len(v)) for k, v in salary.items()])

        if not salaryByName:
            numberByName = dict([(k, 0) for k, v in vacancyNum.items()])
            salaryByName = dict([(k, [0]) for k, v in salary.items()])

        dynamicsSalByYears = self.mean(salary)
        dynamicsSalByYearsForVac = self.mean(salaryByName)
        dynamicsSalByCityDescOrder = self.mean(city)
        numVacByYearForVac = {}

        for y, s in city.items():
            numVacByYearForVac[y] = round(len(s) / count, 4)
        numVacByYearForVac = list(filter(lambda x: x[-1] >= 0.01, [(k, v) for k, v in numVacByYearForVac.items()]))
        numVacByYearForVac.sort(key=lambda x: x[-1], reverse=True)
        dynamicsRateByCity = dict(numVacByYearForVac.copy()[:10])
        numVacByYearForVac = dict(numVacByYearForVac)
        dynamicsSalByCityDescOrder = list(filter(lambda x: x[0] in list(numVacByYearForVac.keys()),
                                                 [(k, v) for k, v in dynamicsSalByCityDescOrder.items()]))
        dynamicsSalByCityDescOrder.sort(key=lambda x: x[-1], reverse=True)
        dynamicsSalByCityDescOrder = dict(dynamicsSalByCityDescOrder[:10])

        return dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrder, dynamicsRateByCity


class InputStart:
    def __init__(self):
        self.filename = 'vacancies_with_skills.csv'
        self.vacancyName = 'C# программист'
        dataset = DataSet(self.filename, self.vacancyName)
        dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrder, dynamicsRateByCity = dataset.getDynamics()

        report = Report(self.vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                        dynamicsSalByCityDescOrder, dynamicsRateByCity)
        report.generate_image()


class Report:
    def __init__(self, vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                 dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10):
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.SalByYears, self.vacancyNum, self.SalByYearsForVac, self.numberByName, self.SalByCityDescOrder, self.RateByCity \
            = dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10

    def generate_excel(self):
        widths = []

        workSheetYear = self.workbook.active

        workSheetYear.title = 'Статистика по годам'
        workSheetYear.append(
            ['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
             'Количество вакансий - ' + self.vacancyName])
        for year in self.SalByYears.keys():
            workSheetYear.append(
                [year, self.SalByYears[year], self.SalByYearsForVac[year], self.vacancyNum[year],
                 self.numberByName[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        for row in data:
            for i, cell in enumerate(row):
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]

        for i, column_width in enumerate(widths, 1):
            workSheetYear.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city1, value1), (city2, value2) in zip(self.SalByCityDescOrder.items(), self.RateByCity.items()):
            data.append([city1, value1, '', city2, value2])
        workSheetCity = self.workbook.create_sheet('Статистика по городам')
        for row in data:
            workSheetCity.append(row)

        widths = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths) > i:
                    if len(cell) > widths[i]:
                        widths[i] = len(cell)
                else:
                    widths += [len(cell)]

        for i, column_width in enumerate(widths, 1):
            workSheetCity.column_dimensions[get_column_letter(i)].width = column_width + 2

        bold = Font(bold=True)

        for index, _ in enumerate(self.SalByCityDescOrder):
            workSheetCity['E' + str(index + 2)].number_format = '0.00%'

        for column in 'ABCDE':
            workSheetYear[column + '1'].font = bold
            workSheetCity[column + '1'].font = bold

        slim = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for column in 'ABDE':
                workSheetCity[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.SalByYears[1] = 1
        for row, _ in enumerate(self.SalByYears):
            for column in 'ABCDE':
                workSheetYear[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('report.xlsx')

    def generate_image(self):

        x = np.arange(len(self.SalByYears.keys()))
        width = 0.35
        areas = []

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(ncols=2, nrows=2)
        fig.patch.set_facecolor('#181d1c')
        fig.patch.set_alpha(0)


        ax1.bar(x - width / 2, self.SalByYears.values(), width, label='средняя з/п')
        ax1.bar(x + width / 2, self.SalByYearsForVac.values(), width, label='з/п C# программист')
        ax1.set_title('Уровень зарплат по годам', color='white')
        ax1.set_xticks(x, self.SalByYears.keys(), rotation=90, color='white')
        ax1.axes.tick_params(color='white', labelcolor='white')
        ax1.legend(fontsize=8)
        ax1.grid(axis='y')

        ax2.bar(x - width / 2, self.vacancyNum.values(), width, label='количество вакансий', ecolor='white')
        ax2.bar(x + width / 2, self.numberByName.values(), width,
                label='количество вакансий\n{0}'.format(self.vacancyName))
        ax2.set_xticks(x, self.vacancyNum.keys(), rotation=90, color='white')
        ax2.set_title('Количество вакансий по годам', color='white')
        ax2.axes.tick_params(color='white', labelcolor='white')
        ax2.legend(fontsize=8)
        ax2.grid(axis='y')
        fig.tight_layout()

        for area in self.SalByCityDescOrder.keys():
            areas.append(str(area).replace(' ', '\n').replace('-', '-\n'))
        y_pos = np.arange(len(areas))
        performance = self.SalByCityDescOrder.values()
        error = np.random.rand(len(areas))
        ax3.barh(y_pos, performance, xerr=error, align='center')
        ax3.set_title('Уровень зарплат по городам', color='white')
        ax3.set_yticks(y_pos, labels=areas, size=6, color='white')
        ax3.axes.tick_params(color='white', labelcolor='white')
        ax3.invert_yaxis()
        ax3.grid(axis='x')

        val = list(self.RateByCity.values()) + [1 - sum(list(self.RateByCity.values()))]
        k = list(self.RateByCity.keys()) + ['Другие']
        colors = plt.get_cmap('Paired_r')(np.linspace(0.2, 0.7, 11))
        ax4.pie(val, labels=k, startangle=170, colors=colors, textprops={'fontsize': 6, 'color': "w"})
        ax4.set_title('Доля вакансий по городам', color='white')
        plt.tight_layout()
        plt.savefig('graph.png', dpi=300)


InputStart()
