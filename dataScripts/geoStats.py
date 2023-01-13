import csv
import math

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

name_list = ["name", "salary_from", "salary_to", "salary_currency", "area_name", "published_at"]

currencyTransfer = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class Vacancy:
    def __init__(self, vacancies):
        self.name = vacancies[name_list[0]]
        self.averageSalary = math.floor((float(vacancies[name_list[1]]) + float(vacancies[name_list[2]])) / 2) * \
                             currencyTransfer[vacancies[name_list[3]]]
        self.areaName = vacancies[name_list[4]]
        self.publicationYear = int(vacancies[name_list[5]][:4])


class DataSet:
    def __init__(self, filename, vacancyName):
        self.filename = filename
        self.vacancyName = vacancyName

    def csv_reader_raw(self):
        with open(self.filename, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            headerLength = len(header)
            for row in reader:
                if '' not in row and len(row) == headerLength:
                    yield dict(zip(header, row))

    @staticmethod
    def increment(dictIN, k, c):
        if k in dictIN:
            dictIN[k] += c
        else:
            dictIN[k] = c

    @staticmethod
    def mean(dictAV):
        newDict = {}
        for k, v in dictAV.items():
            newDict[k] = int(sum(v) / len(v))
        return newDict

    def getDynamics(self):
        salary = {}
        salaryByName = {}
        city = {}
        count = 0

        for vacancyDictionary in self.csv_reader_raw():
            vacancy = Vacancy(vacancyDictionary)
            self.increment(salary, vacancy.publicationYear, [vacancy.averageSalary])
            if any(x.lower() in ['с#', 'c sharp', 'шарп', 'c#'] for x in vacancy.name.split(' ')):
                self.increment(salaryByName, vacancy.publicationYear, [vacancy.averageSalary])
            self.increment(city, vacancy.areaName, [vacancy.averageSalary])
            count += 1

        numberByName = dict([(k, len(v)) for k, v in salaryByName.items()])
        vacancyNum = dict([(k, len(v)) for k, v in salary.items()])

        if not salaryByName:
            numberByName = dict([(k, 0) for k, v in vacancyNum.items()])
            salaryByName = dict([(k, [0]) for k, v in salary.items()])

        dynamicsSalByYears = self.mean(salary)
        dynamicsSalByYearsForVac = self.mean(salaryByName)
        dynamicsSalByCityDescOrder = self.mean(city)
        numVacByYearForVac = {}

        for y, s in city.items():
            numVacByYearForVac[y] = round(len(s) / count, 4)
        numVacByYearForVac = list(filter(lambda x: x[-1] >= 0.01, [(k, v) for k, v in numVacByYearForVac.items()]))
        numVacByYearForVac.sort(key=lambda x: x[-1], reverse=True)
        dynamicsRateByCity = dict(numVacByYearForVac.copy()[:10])
        numVacByYearForVac = dict(numVacByYearForVac)
        dynamicsSalByCityDescOrder = list(filter(lambda x: x[0] in list(numVacByYearForVac.keys()),
                                                 [(k, v) for k, v in dynamicsSalByCityDescOrder.items()]))
        dynamicsSalByCityDescOrder.sort(key=lambda x: x[-1], reverse=True)
        dynamicsSalByCityDescOrder = dict(dynamicsSalByCityDescOrder[:10])

        return dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrder, dynamicsRateByCity


class InputStart:
    def __init__(self):
        # self.filename = input('Введите название файла: ')
        # self.vacancyName = input('Введите название профессии: ')
        self.filename = 'vacancies_with_skills.csv'
        self.vacancyName = 'C# Програмист'
        dataset = DataSet(self.filename, self.vacancyName)
        dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrder, dynamicsRateByCity = dataset.getDynamics()

        self.report = Report(self.vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                             dynamicsSalByCityDescOrder, dynamicsRateByCity)
        self.report.generateExelFromSheets()


class Report:
    def __init__(self, vacancyName, dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName,
                 dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10):
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.SalByYears, self.vacancyNum, self.SalByYearsForVac, self.numberByName, self.SalByCityDescOrder, self.RateByCity \
            = dynamicsSalByYears, vacancyNum, dynamicsSalByYearsForVac, numberByName, dynamicsSalByCityDescOrderTop10, dynamicsRateByCityTop10

    def getYearSheet(self):

        work_sheet1 = self.workbook.active
        work_sheet1.title = 'Статистика по годам'
        work_sheet1.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
                            'Количество вакансий - ' + self.vacancyName])
        for year in self.SalByYears.keys():
            work_sheet1.append(
                [year, self.SalByYears[year], self.SalByYearsForVac[year], self.vacancyNum[year],
                 self.numberByName[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            work_sheet1.column_dimensions[get_column_letter(i)].width = column_width + 2
        return work_sheet1

    def getCitySheet(self):

        new_data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (city1, value1), (city2, value2) in zip(self.SalByCityDescOrder.items(), self.RateByCity.items()):
            new_data.append([city1, value1, '', city2, value2])
        work_sheet2 = self.workbook.create_sheet('Статистика по городам')
        for r in new_data:
            work_sheet2.append(r)

        column_widths = []
        for r in new_data:
            for i, cell in enumerate(r):
                cell = str(cell)
                if len(column_widths) > i:
                    if len(cell) > column_widths[i]:
                        column_widths[i] = len(cell)
                else:
                    column_widths += [len(cell)]

        for i, column_width in enumerate(column_widths, 1):
            work_sheet2.column_dimensions[get_column_letter(i)].width = column_width + 2

        return work_sheet2, len(new_data)

    def generateExelFromSheets(self):
        yearSheet = self.getYearSheet()
        citySheet, len_new_data = self.getCitySheet()
        bold = Font(bold=True)
        for c in 'ABCDE':
            yearSheet[c + '1'].font = bold
            citySheet[c + '1'].font = bold

        for index, _ in enumerate(self.SalByCityDescOrder):
            citySheet['E' + str(index + 2)].number_format = '0.00%'

        slim = Side(border_style='thin', color='00000000')

        for row in range(len_new_data):
            for column in 'ABDE':
                citySheet[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.SalByYears[1] = 1
        for row, _ in enumerate(self.SalByYears):
            for column in 'ABCDE':
                yearSheet[column + str(row + 1)].border = Border(left=slim, bottom=slim, right=slim, top=slim)

        self.workbook.save('GeoReport.xlsx')


InputStart()
